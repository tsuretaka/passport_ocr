import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook

def init_excel(file_path):
    """Excelファイルが存在しない場合、ヘッダー付きで作成する"""
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Passport Data"
        headers = [
            "登録日時", "旅券番号", "氏名(姓)", "氏名(名)", 
            "生年月日", "性別", "国籍", "本籍", "発行年月日", "有効期間満了日", 
            "住所(手入力)", "備考", "画像ファイル名"
        ]
        ws.append(headers)
        # 親ディレクトリ作成
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)

def load_data_as_df(file_path):
    """ExcelデータをDataFrameとして読み込む"""
    if not os.path.exists(file_path):
        return pd.DataFrame()
    df = pd.read_excel(file_path)
    return df.fillna("")

def ensure_excel_columns(file_path):
    """
    既存のExcelファイルを確認し、新しい列（本籍、発行年月日）がなければ追加し、
    列の順序を最新の定義に合わせるマイグレーション処理を行う。
    """
    if not os.path.exists(file_path):
        return

    try:
        df = pd.read_excel(file_path)
        
        # 最新の定義ヘッダー
        expected_headers = [
            "登録日時", "旅券番号", "氏名(姓)", "氏名(名)", 
            "生年月日", "性別", "国籍", "本籍", "発行年月日", "有効期間満了日", 
            "住所(手入力)", "備考", "画像ファイル名"
        ]
        
        changed = False
        # 足りない列を追加
        for col in expected_headers:
            if col not in df.columns:
                df[col] = "" # 空文字で埋める
                changed = True
        
        # 不要な列があるか、順序が違う場合も修正が必要だが、
        # 基本的には expected_headers に含まれる列だけを抽出して並べ替える
        # (既存の未知の列は消さない方が安全だが、今回は定義通りに整頓する)
        
        if changed or list(df.columns) != expected_headers:
            # 存在しない列は除外せず、新しい定義にある列だけを並べる（他は残す）
            # 新しい順序で再構築
            new_df = df.reindex(columns=expected_headers)
            new_df.to_excel(file_path, index=False)
            
    except Exception as e:
        print(f"Migration failed: {e}")

# save_passport_data の前にマイグレーションを実行するように変更
def save_passport_data(file_path, data, image_filename=""):
    """
    パスポートデータをExcelに追記する。
    data: dict (ocr_utils.parse_mrz の戻り値 + 住所など)
    """
    if os.path.exists(file_path):
        ensure_excel_columns(file_path)
    else:
        init_excel(file_path)
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        data.get("passport_no", ""),
        data.get("surname", ""),
        data.get("given_name", ""),
        data.get("birth_date", ""),
        data.get("sex", ""),
        data.get("nationality", ""),
        data.get("domicile", ""),
        data.get("issue_date", ""),
        data.get("expiry_date", ""),
        data.get("address", ""),
        data.get("note", ""),
        image_filename
    ]
    
    ws.append(row)
    wb.save(file_path)

def save_all_data(file_path, df):
    """
    DataFrameの内容でExcelファイルを上書き保存する（削除・編集反映用）
    """
    if df is None: return
    df.to_excel(file_path, index=False)
